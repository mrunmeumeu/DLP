#the file is named as text14.exe as of testing
import win32clipboard
import os
from pathlib import Path
import re
import pyperclip
import time
import zipfile
import tarfile
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import tempfile
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
import pystray
from PIL import Image, ImageDraw
import threading
import tkinter as tk
from tkinter import simpledialog, messagebox, Listbox, END
import firebase_admin
from firebase_admin import credentials, db
import json
import sys
# Define sensitive keywords
# cred = credentials.Certificate("C:\\Program Files\\DLP\\clipboard-81621-firebase-adminsdk-rpim7-dd58d299af.json")

# # Initialize Firebase Admin SDK
# firebase_admin.initialize_app(cred, {
#     'databaseURL': "https://clipboard-81621-default-rtdb.asia-southeast1.firebasedatabase.app"  # Replace with your Firebase Database URL
# })

FIREBASE_CONFIG_PATH = r"C:\\Program Files\\DLP\\firebase_config.json"

program_files_dir = "C:\\Program Files"  # Adjust the base directory if needed
KEYWORDS_FILE_PATH = os.path.join(program_files_dir, "DLP", "keywords.txt")

def load_firebase_config():
    try:
        with open(FIREBASE_CONFIG_PATH, 'r') as config_file:
            config = json.load(config_file)
            return config
    except Exception as e:
        print(f"Failed to load Firebase configuration: {e}")
        sys.exit()

# Initialize Firebase
def initialize_firebase():
    config = load_firebase_config()
    if not firebase_admin._apps:  # Avoid re-initializing
        try:
            cred = credentials.Certificate(config["serviceAccountPath"])
            firebase_admin.initialize_app(cred, {'databaseURL': config["databaseURL"]})
            print("Firebase initialized successfully.")
        except Exception as e:
            print(f"Failed to initialize Firebase: {e}")
            sys.exit()

initialize_firebase()

def load_sensitive_keywords():
    if os.path.exists(KEYWORDS_FILE_PATH):
        with open(KEYWORDS_FILE_PATH, 'r', encoding='utf-8') as file:
            return [line.strip() for line in file if line.strip()]  # Read non-empty lines
    return []

# Function to save sensitive keywords to a file
def save_sensitive_keywords(keywords):
    try:
        with open(KEYWORDS_FILE_PATH, 'w', encoding='utf-8') as file:
            for keyword in keywords:
                file.write(f"{keyword}\n")
    except Exception as e:
        print(f"Error saving keywords to file: {e}")

# Ensure the keywords file exists
if not os.path.exists(KEYWORDS_FILE_PATH):
    os.makedirs(os.path.dirname(KEYWORDS_FILE_PATH), exist_ok=True)
    with open(KEYWORDS_FILE_PATH, 'w', encoding='utf-8') as file:
        file.write("")  # Create an empty file

# Initialize sensitive keywords
SENSITIVE_KEYWORDS = load_sensitive_keywords()
keywords_file_last_modified = os.path.getmtime(KEYWORDS_FILE_PATH)

def check_for_keywords_update():
    global SENSITIVE_KEYWORDS, keywords_file_last_modified
    try:
        current_modified_time = os.path.getmtime(KEYWORDS_FILE_PATH)
        if current_modified_time != keywords_file_last_modified:
            print("Keywords file updated. Reloading keywords...")
            SENSITIVE_KEYWORDS = load_sensitive_keywords()
            keywords_file_last_modified = current_modified_time
            print(f"Updated sensitive keywords: {SENSITIVE_KEYWORDS}")
    except Exception as e:
        print(f"Error checking for keywords update: {e}")


class KeywordsFileEventHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path == KEYWORDS_FILE_PATH:
            print("Keywords file updated by admin. Reloading keywords...")
            check_for_keywords_update()  # Reload the keywords list

def start_keywords_file_watcher():
    """
    Starts a file watcher on the keywords file to detect modifications.
    """
    event_handler = KeywordsFileEventHandler()
    observer = Observer()
    observer.schedule(event_handler, path=os.path.dirname(KEYWORDS_FILE_PATH), recursive=False)
    observer.start()
    return observer


def get_device_username():
    try:
        return os.getlogin()  # This returns the currently logged-in user
    except Exception as e:
        print(f"Error getting username: {e}")
        return "Unknown"
# Function to check if a string contains sensitive keywords
def contains_sensitive_keywords(text):
    for keyword in SENSITIVE_KEYWORDS:
        if re.search(rf'\b{keyword}\b', text, re.IGNORECASE):
            log_event("Sensitive content detected and clipboard cleared.", detected_word=keyword)
            return True
    return False

# Function to get clipboard content as a file path if files are copied
def get_clipboard_files():
    attempts = 0
    while attempts < 5:
        try:
            win32clipboard.OpenClipboard()
            if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_HDROP):
                file_paths = win32clipboard.GetClipboardData(win32clipboard.CF_HDROP)
                return list(file_paths)
        except Exception as e:
            print(f"Error accessing clipboard (attempt {attempts + 1}): {e}")
            attempts += 1
            time.sleep(1)  # Wait before retrying clipboard access
        finally:
            try:
                win32clipboard.CloseClipboard()
            except Exception as close_error:
                print(f"Error closing clipboard: {close_error}")
    return None

# Function to extract zip files
def extract_zip(file_path, extract_to):
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
    except Exception as e:
        print(f"Error extracting zip file {file_path}: {e}")

# Function to extract tar or tar.gz files
def extract_tar(file_path, extract_to):
    try:
        with tarfile.open(file_path, 'r:*') as tar_ref:
            tar_ref.extractall(extract_to)
    except Exception as e:
        print(f"Error extracting tar file {file_path}: {e}")

# Function to read the content of a text file
def read_file_content(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
    return None

# Function to clear the clipboard using pyperclip
def log_event(event_description, detected_word=None):
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    username = get_device_username()
    # Create a reference to your logs in the database
    ref = db.reference('logs')

    # Prepare the data to log
    log_data = {
        'timestamp': timestamp,
        'event_description': event_description,
        'username': username 
    }

    # Include the detected word if provided
    if detected_word:
        log_data['detected_word'] = detected_word

    # Push the new log to Firebase
    new_log = ref.push(log_data)

    print(f"Log added: {new_log.key} - {event_description}")
    if detected_word:
        print(f"Detected word: {detected_word}")

def clear_clipboard():
    attempts = 0
    while attempts < 5:
        try:
            pyperclip.copy('')  # Replace clipboard content with an empty string using pyperclip
            print("Clipboard cleared.")
            # log_event("Sensitive content detected and clipboard cleared.")
            return
        except Exception as e:
            print(f"Error clearing clipboard (attempt {attempts + 1}): {e}")
            attempts += 1
            time.sleep(1)  # Wait before retrying clipboard access

# Function to scan the contents of extracted files (recursive)
def scan_extracted_files(extracted_dir):
    for root, _, files in os.walk(extracted_dir):
        for file in files:
            file_path = os.path.join(root, file)
            file_name = Path(file_path).name

            # Check if the file name contains sensitive keywords
            if contains_sensitive_keywords(file_name):
                print(f"Sensitive file name detected: {file_name}")
                return True

            # Check if the file is a supported file type
            if scan_regular_file(file_path):
                return True

    return False

# Function to scan regular files, including PDFs, Word, and Excel files
def scan_regular_file(file_path):
    file_name = Path(file_path).name

    # Check if the file name contains sensitive keywords
    if contains_sensitive_keywords(file_name):
        print(f"Sensitive file name detected: {file_name}")
        return True

    # Check content based on file extension
    if file_path.endswith('.txt'):
        file_content = read_file_content(file_path)
        if file_content and contains_sensitive_keywords(file_content):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Check for PDF files
    elif file_path.endswith('.pdf'):
        if scan_pdf_file(file_path):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Check for Word files (.docx)
    elif file_path.endswith('.docx'):
        if scan_word_file(file_path):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Check for Excel files (.xlsx)
    elif file_path.endswith('.xlsx'):
        if scan_excel_file(file_path):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Handle nested compressed files (zip or tar)
    elif file_path.endswith('.zip') or file_path.endswith('.tar.gz') or file_path.endswith('.tar'):
        # Create a temporary directory to extract the compressed file
        with tempfile.TemporaryDirectory() as temp_dir:
            if file_path.endswith('.zip'):
                extract_zip(file_path, temp_dir)
            elif file_path.endswith('.tar') or file_path.endswith('.tar.gz'):
                extract_tar(file_path, temp_dir)
            # Recursively scan the extracted files
            if scan_extracted_files(temp_dir):
                return True

    return False

# Function to scan a PDF file for sensitive content
def scan_pdf_file(file_path):
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return contains_sensitive_keywords(text)
    except Exception as e:
        print(f"Error reading PDF file {file_path}: {e}")
    return False

# Function to scan a Word file for sensitive content (only for .docx)
def scan_word_file(file_path):
    try:
        doc = Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs])
        return contains_sensitive_keywords(text)
    except Exception as e:
        print(f"Error reading Word file {file_path}: {e}")
    return False

# Function to scan an Excel file for sensitive content
def scan_excel_file(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        text = ""

        # Iterate through all sheets and cells to collect the content
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_data = " ".join([str(cell) for cell in row if cell is not None])
                text += row_data + " "

        # Check for sensitive keywords in the collected content
        return contains_sensitive_keywords(text)
    except Exception as e:
        print(f"Error reading Excel file {file_path}: {e}")
    return False

# Main function to monitor clipboard and scan files (regular or compressed)
def monitor_clipboard():
    last_clipboard_content = None

    while True:
        check_for_keywords_update() 
        clipboard_files = get_clipboard_files()

        if clipboard_files and clipboard_files != last_clipboard_content:
            for file_path in clipboard_files:
                file_name = Path(file_path).name

                # If it's a regular file or compressed archive, scan it
                if scan_regular_file(file_path):
                    print(f"Sensitive content found in {file_name}. Clearing clipboard.")
                    clear_clipboard()
                    break

            last_clipboard_content = clipboard_files

        time.sleep(1)

# Function to create a tray icon
def create_tray_icon():
    # Create an icon for the system tray
    image = Image.new('RGB', (64, 64), color=(73, 109, 137))
    draw = ImageDraw.Draw(image)
    draw.text((10, 10), "CM", fill=(255, 255, 255))  # "CM" for Clipboard Monitor

    def on_quit(icon, item):
        icon.stop()

    # Function to show the admin GUI
    def show_gui(icon, item):
        create_gui()

    icon = pystray.Icon("clipboard_monitor", image, "Clipboard Monitor", menu=pystray.Menu(
        pystray.MenuItem("Show Admin", show_gui),
        pystray.MenuItem("Quit", on_quit)
    ))

    icon.run()

# Function to create a simple Tkinter window with admin access button
def create_gui():
    root = tk.Tk()
    root.title("Clipboard Monitor")

    # Admin button to modify sensitive keywords
    admin_button = tk.Button(root, text="Admin Access (Edit Keywords)", command=admin_access)
    admin_button.pack(pady=20)

    root.mainloop()

# Function to handle admin access and keyword management
def admin_access():
    password = "tinker@tl"  # Hardcoded password for simplicity (you can change this)

    # Get password from the user using a prompt
    entered_password = simpledialog.askstring("Admin Access", "Enter Admin Password:", show='*')

    if entered_password == password:
        # Admin authenticated
        admin_gui()
    else:
        messagebox.showerror("Access Denied", "Incorrect password.")

# Function to open the admin GUI to manage keywords
def admin_gui():
    global SENSITIVE_KEYWORDS

    # Create a new window
    admin_window = tk.Toplevel()
    admin_window.title("Manage Sensitive Keywords")

    # Create a Listbox to display current keywords
    keyword_listbox = Listbox(admin_window, height=10, width=50)
    keyword_listbox.pack(pady=10)

    # Add current keywords to the Listbox
    for keyword in SENSITIVE_KEYWORDS:
        keyword_listbox.insert(END, keyword)

    # Function to add a new keyword
    def add_keyword():
        new_keyword = simpledialog.askstring("Add Keyword", "Enter a new keyword:")
        if new_keyword and new_keyword.strip():
            new_keyword = new_keyword.strip()
            if new_keyword not in SENSITIVE_KEYWORDS:
                SENSITIVE_KEYWORDS.append(new_keyword)
                keyword_listbox.insert(END, new_keyword)
            else:
                messagebox.showwarning("Warning", "Keyword already exists!")

    # Function to remove the selected keyword
    def remove_keyword():
        selected_keyword = keyword_listbox.curselection()
        if selected_keyword:
            keyword = keyword_listbox.get(selected_keyword)
            SENSITIVE_KEYWORDS.remove(keyword)
            keyword_listbox.delete(selected_keyword)
        else:
            messagebox.showwarning("Warning", "Please select a keyword to remove.")

    # Function to clear all keywords
    def clear_keywords():
        confirm = messagebox.askyesno("Confirm", "Are you sure you want to clear all keywords?")
        if confirm:
            SENSITIVE_KEYWORDS.clear()
            keyword_listbox.delete(0, END)

    # Function to apply changes and close the window
    def apply_changes():
        admin_window.destroy()
        messagebox.showinfo("Success", "Changes have been applied successfully!")

    # Add buttons to add, remove, clear, and apply changes
    tk.Button(admin_window, text="Add Keyword", command=add_keyword).pack(pady=5)
    tk.Button(admin_window, text="Remove Keyword", command=remove_keyword).pack(pady=5)
    tk.Button(admin_window, text="Clear All Keywords", command=clear_keywords).pack(pady=5)
    tk.Button(admin_window, text="Apply and Close", command=apply_changes).pack(pady=10)

if __name__ == "__main__":
    # Start the keywords file watcher
    observer = start_keywords_file_watcher()

    # Start clipboard monitor in a background thread
    monitor_thread = threading.Thread(target=monitor_clipboard, daemon=True)
    monitor_thread.start()

    # Start tray icon in a separate thread
    tray_thread = threading.Thread(target=create_tray_icon, daemon=True)
    tray_thread.start()

    try:
        # Keep the main thread alive
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
