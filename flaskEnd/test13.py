import win32clipboard
import os
from pathlib import Path
import re
import pyperclip
import time
import zipfile
import tarfile
import tempfile
from PyPDF2 import PdfReader  # For PDF files
from docx import Document     # For Word files (.docx)
from openpyxl import load_workbook  # For Excel files (.xlsx)
import pystray
from PIL import Image, ImageDraw
import threading
import tkinter as tk
from tkinter import simpledialog, messagebox, Listbox, END

# Path to store keywords persistently
KEYWORD_FILE_PATH = Path.home() / "Desktop" /"DLP"/ "keywords.txt"

# Ensure the directory exists
KEYWORD_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
# Load sensitive keywords from a file
def load_keywords():
    if os.path.exists(KEYWORD_FILE_PATH):
        print("Loading keywords from file...")
        with open(KEYWORD_FILE_PATH, "r") as f:
            keywords = [line.strip() for line in f if line.strip()]
        print("Keywords loaded:", keywords)
        return keywords
    print("Keyword file not found. Using default keywords.")
    return ["tl", "comission", "confidential", "kickback"]  # Default keywords if no file exists

# Save sensitive keywords to a file
def save_keywords():
    with open(KEYWORD_FILE_PATH, "w") as f:
        for keyword in SENSITIVE_KEYWORDS:
            f.write(keyword + "\n")

# Initialize sensitive keywords
SENSITIVE_KEYWORDS = load_keywords()

# Function to reload keywords from file periodically
def reload_keywords():
    global SENSITIVE_KEYWORDS
    SENSITIVE_KEYWORDS = load_keywords()
    print("reloaded")

# Function to check if a string contains sensitive keywords
def contains_sensitive_keywords(text):
    for keyword in SENSITIVE_KEYWORDS:
        if re.search(rf'\b{keyword}\b', text, re.IGNORECASE):
            return True
    return False

# Function to get clipboard content as a file path if files are copied
def get_clipboard_files():
    attempts = 0
    while attempts < 5:
        try:
            win32clipboard.OpenClipboard()
            if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_HDROP):
                file_paths = win32clipboard.GetClipboardData(win32clipboard.CF_HDROP)
                return list(file_paths)
        except Exception as e:
            print(f"Error accessing clipboard (attempt {attempts + 1}): {e}")
            attempts += 1
            time.sleep(1)  # Wait before retrying clipboard access
        finally:
            try:
                win32clipboard.CloseClipboard()
            except Exception as close_error:
                print(f"Error closing clipboard: {close_error}")
        time.sleep(1)
    return None

# Function to extract zip files
def extract_zip(file_path, extract_to):
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
    except Exception as e:
        print(f"Error extracting zip file {file_path}: {e}")

# Function to extract tar or tar.gz files
def extract_tar(file_path, extract_to):
    try:
        with tarfile.open(file_path, 'r:*') as tar_ref:
            tar_ref.extractall(extract_to)
    except Exception as e:
        print(f"Error extracting tar file {file_path}: {e}")

# Function to read the content of a text file
def read_file_content(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
    return None

# Function to clear the clipboard using pyperclip
def clear_clipboard():
    attempts = 0
    while attempts < 3:
        try:
            pyperclip.copy('')  # Replace clipboard content with an empty string using pyperclip
            print("Clipboard cleared.")
            return
        except Exception as e:
            print(f"Error clearing clipboard (attempt {attempts + 1}): {e}")
            attempts += 1
        time.sleep(2)  # Wait before retrying clipboard access

# Function to scan the contents of extracted files (recursive)
def scan_extracted_files(extracted_dir):
    for root, _, files in os.walk(extracted_dir):
        for file in files:
            file_path = os.path.join(root, file)
            file_name = Path(file_path).name

            # Check if the file name contains sensitive keywords
            if contains_sensitive_keywords(file_name):
                print(f"Sensitive file name detected: {file_name}")
                return True

            # Check if the file is a supported file type
            if scan_regular_file(file_path):
                return True

    return False

# Function to scan regular files, including PDFs, Word, and Excel files
def scan_regular_file(file_path):
    file_name = Path(file_path).name

    # Check if the file name contains sensitive keywords
    if contains_sensitive_keywords(file_name):
        print(f"Sensitive file name detected: {file_name}")
        return True

    # Check content based on file extension
    if file_path.endswith('.txt'):
        file_content = read_file_content(file_path)
        if file_content and contains_sensitive_keywords(file_content):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Check for PDF files
    elif file_path.endswith('.pdf'):
        if scan_pdf_file(file_path):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Check for Word files (.docx)
    elif file_path.endswith('.docx'):
        if scan_word_file(file_path):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Check for Excel files (.xlsx)
    elif file_path.endswith('.xlsx'):
        if scan_excel_file(file_path):
            print(f"Sensitive content detected in {file_name}")
            return True

    # Handle nested compressed files (zip or tar)
    elif file_path.endswith('.zip') or file_path.endswith('.tar.gz') or file_path.endswith('.tar'):
        # Create a temporary directory to extract the compressed file
        with tempfile.TemporaryDirectory() as temp_dir:
            if file_path.endswith('.zip'):
                extract_zip(file_path, temp_dir)
            elif file_path.endswith('.tar') or file_path.endswith('.tar.gz'):
                extract_tar(file_path, temp_dir)
            # Recursively scan the extracted files
            if scan_extracted_files(temp_dir):
                return True

    return False

# Function to scan a PDF file for sensitive content
def scan_pdf_file(file_path):
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return contains_sensitive_keywords(text)
    except Exception as e:
        print(f"Error reading PDF file {file_path}: {e}")
    return False

# Function to scan a Word file for sensitive content (only for .docx)
def scan_word_file(file_path):
    try:
        doc = Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs])
        return contains_sensitive_keywords(text)
    except Exception as e:
        print(f"Error reading Word file {file_path}: {e}")
    return False

# Function to scan an Excel file for sensitive content
def scan_excel_file(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        text = ""

        # Iterate through all sheets and cells to collect the content
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_data = " ".join([str(cell) for cell in row if cell is not None])
                text += row_data + " "

        # Check for sensitive keywords in the collected content
        return contains_sensitive_keywords(text)
    except Exception as e:
        print(f"Error reading Excel file {file_path}: {e}")
    return False

# Main function to monitor clipboard and scan files (regular or compressed)
def monitor_clipboard():
    last_clipboard_content = None
    last_reload_time = time.time()

    while True:
        # Reload keywords every 10 seconds to check for updates
        if time.time() - last_reload_time > 10:
            print("reloading")
            reload_keywords()
            last_reload_time = time.time()

        clipboard_files = get_clipboard_files()

        if clipboard_files and clipboard_files != last_clipboard_content:
            for file_path in clipboard_files:
                file_name = Path(file_path).name

                # If it's a regular file or compressed archive, scan it
                if scan_regular_file(file_path):
                    print(f"Sensitive content found in {file_name}. Clearing clipboard.")
                    clear_clipboard()
                    break

            last_clipboard_content = clipboard_files

        time.sleep(1)

# Function to create a tray icon
def create_tray_icon():
    # Create an icon for the system tray
    image = Image.new('RGB', (64, 64), color=(73, 109, 137))
    draw = ImageDraw.Draw(image)
    draw.text((10, 10), "CM", fill=(255, 255, 255))  # "CM" for Clipboard Monitor

    def on_quit(icon, item):
        icon.stop()

    # Function to show the admin GUI
    def show_gui(icon, item):
        create_gui()

    icon = pystray.Icon("clipboard_monitor", image, "Clipboard Monitor", menu=pystray.Menu(
        pystray.MenuItem("Show Admin", show_gui),
        pystray.MenuItem("Quit", on_quit)
    ))

    icon.run()

# Function to create a simple Tkinter window with admin access button
def create_gui():
    root = tk.Tk()
    root.title("Clipboard Monitor")

    # Admin button to modify sensitive keywords
    admin_button = tk.Button(root, text="Admin Access (Edit Keywords)", command=admin_access)
    admin_button.pack(pady=20)

    root.mainloop()

# Function to handle admin access and keyword management
def admin_access():
    password = "tinker@tl"  # Hardcoded password for simplicity (you can change this)

    # Get password from the user using a prompt
    entered_password = simpledialog.askstring("Admin Access", "Enter Admin Password:", show='*')

    if entered_password == password:
        # Admin authenticated
        admin_gui()
    else:
        messagebox.showerror("Access Denied", "Incorrect password.")

# Function to open the admin GUI to manage keywords
def admin_gui():
    global SENSITIVE_KEYWORDS

    # Create a new window
    admin_window = tk.Toplevel()
    admin_window.title("Manage Sensitive Keywords")

    # Create a Listbox to display current keywords
    keyword_listbox = Listbox(admin_window, height=10, width=50)
    keyword_listbox.pack(pady=10)

    # Add current keywords to the Listbox
    for keyword in SENSITIVE_KEYWORDS:
        keyword_listbox.insert(END, keyword)

    # Function to add a new keyword
    def add_keyword():
        new_keyword = simpledialog.askstring("Add Keyword", "Enter a new keyword:")
        if new_keyword and new_keyword.strip():
            new_keyword = new_keyword.strip()
            if new_keyword not in SENSITIVE_KEYWORDS:
                SENSITIVE_KEYWORDS.append(new_keyword)
                keyword_listbox.insert(END, new_keyword)
                save_keywords()  # Save to file immediately
            else:
                messagebox.showwarning("Warning", "Keyword already exists!")

    # Function to remove the selected keyword
    def remove_keyword():
        selected_keyword = keyword_listbox.curselection()
        if selected_keyword:
            keyword = keyword_listbox.get(selected_keyword)
            SENSITIVE_KEYWORDS.remove(keyword)
            keyword_listbox.delete(selected_keyword)
            save_keywords()  # Save to file immediately
        else:
            messagebox.showwarning("Warning", "Please select a keyword to remove.")

    # Function to clear all keywords
    def clear_keywords():
        confirm = messagebox.askyesno("Confirm", "Are you sure you want to clear all keywords?")
        if confirm:
            SENSITIVE_KEYWORDS.clear()
            keyword_listbox.delete(0, END)
            save_keywords()  # Save to file immediately

    # Function to apply changes and close the window
    def apply_changes():
        admin_window.destroy()
        messagebox.showinfo("Success", "Changes have been applied successfully!")

    # Add buttons to add, remove, clear, and apply changes
    tk.Button(admin_window, text="Add Keyword", command=add_keyword).pack(pady=5)
    tk.Button(admin_window, text="Remove Keyword", command=remove_keyword).pack(pady=5)
    tk.Button(admin_window, text="Clear All Keywords", command=clear_keywords).pack(pady=5)
    tk.Button(admin_window, text="Apply and Close", command=apply_changes).pack(pady=10)

if __name__ == "__main__":
    # Start clipboard monitor in a background thread
    monitor_thread = threading.Thread(target=monitor_clipboard, daemon=True)
    monitor_thread.start()

    # Start tray icon in a separate thread
    tray_thread = threading.Thread(target=create_tray_icon, daemon=True)
    tray_thread.start()

    # Keep the main thread alive
    while True:
        time.sleep(1)
